# Importazione di tutti i vari moduli utilizzati per questo file.
import csv
import cv2
import openpyxl
import os
import numpy as np
import mediapipe as mp
import pandas as pd
import argparse
import shutil
import matplotlib.pyplot as plt
from keras.models import Sequential
from keras.regularizers import l2
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import f1_score, confusion_matrix
from keras.layers import TimeDistributed, Conv2D, MaxPooling2D, Flatten, LSTM, Dense, BatchNormalization
from keras.callbacks import EarlyStopping, TensorBoard, ReduceLROnPlateau
from sklearn.utils import compute_sample_weight
from imblearn.over_sampling import RandomOverSampler
from keras.optimizers import Adam, RMSprop

from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.model_selection import train_test_split
from sklearn.model_selection import cross_val_score
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix


"""
    Funzione **calcola_media_colonne** per calcolare la media delle colonne in un file Excel.
    Prende in input il percorso del file Excel.
    Restituisce una lista di tuple, ognuna contenente il nome del video e l'etichetta corrispondente.
"""



def calcola_media_colonne(file_excel):
    wb = openpyxl.load_workbook(file_excel)
    foglio = wb.worksheets[0]
    media_colonne = []
    etichette = []
    nomi_video = []

    # Ottenere etichette
    for riga in foglio.iter_rows(min_row=2, max_row=2, min_col=2):
        for cella in riga:
            if isinstance(cella.value, str):
                if cella.value.startswith("The person in the video appears"):
                    etichetta = "-" + cella.value.split("-")[-1].strip()
                else:
                    etichetta = cella.value.strip()
                etichette.append(etichetta)
            else:
                etichette.append('')

    # Ottenere nomi dei video
    for cella in foglio[1][1:]:
        if isinstance(cella.value, str):
            nomi_video.append(cella.value)
        else:
            nomi_video.append('')

    # Calcolo media delle colonne
    for colonna in foglio.iter_cols(min_row=3, min_col=2):
        valori_colonna = [cella.value for cella in colonna if isinstance(cella.value, (int, float))]
        if valori_colonna:
            media = sum(valori_colonna) / (foglio.max_row - 2)
        else:
            media = 0
        media_colonne.append(media)

    # Calcolo delle percentuali per ciascun video
    risultato = []
    ultimo_nome_video = ''
    i = 0
    while i < len(nomi_video):
        sub_array = media_colonne[i:i + 4]
        sub_etichette = etichette[i:i + 4]
        nome_video = nomi_video[i]

        somma_emozioni = sum(sub_array)
        if somma_emozioni > 0:
            percentuali = [(val / somma_emozioni) for val in sub_array]
        else:
            percentuali = [0] * 4

        for j in range(4):
            risultato.append((nome_video, sub_etichette[j], f"{percentuali[j]:.5f}"))

        i += 4

    return risultato


"""
   Funzione **apply_landmarks** per applicare i landmarks a un frame.
   Prende in input il frame, l'oggetto mediapipe, la larghezza e l'altezza desiderate.
   Restituisce il frame con i landmarks applicati.
"""


def apply_landmarks(frame, mp_holistic, target_width, target_height):
    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    frame_rgb = cv2.GaussianBlur(frame_rgb, (5, 5), 0)
    frame_rgb = cv2.resize(frame_rgb, (target_width, target_height))

    results = mp_holistic.process(frame_rgb)
    frame_with_landmarks = frame_rgb.copy()
    mp.solutions.drawing_utils.draw_landmarks(
        frame_with_landmarks, results.pose_landmarks, mp.solutions.holistic.POSE_CONNECTIONS)
    mp.solutions.drawing_utils.draw_landmarks(
        frame_with_landmarks, results.face_landmarks)

    return frame_with_landmarks


"""
    Funzione **process_video_folder** per processare una cartella di video.
    Prende in input il percorso della cartella di input, il percorso della cartella di output, la larghezza e l'altezza desiderate.
    Restituisce una lista di tuple, ognuna contenente il nome del video e la lista dei frame con i landmarks applicati.
"""


def process_video_folder(input_folder, output_folder, target_width, target_height):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    mp_holistic = mp.solutions.holistic.Holistic()
    processed_videos = []

    for filename in os.listdir(input_folder):
        if filename.lower().endswith(('.avi', '.mp4', '.mkv', '.mov')):
            video_path = os.path.join(input_folder, filename)
            video_name = os.path.splitext(filename)[0]
            output_subfolder = os.path.join(output_folder, video_name)
            index = 1
            while os.path.exists(output_subfolder):
                output_subfolder = os.path.join(output_folder, f"{video_name}_{index}")
                index += 1
            os.makedirs(output_subfolder)

            cap = cv2.VideoCapture(video_path)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_with_landmarks_list = []
            frame_count = 0
            start_time = 0
            end_time = start_time + 1

            while True:
                ret, frame = cap.read()
                if not ret:
                    break

                current_time = frame_count / fps

                if current_time >= start_time and current_time < end_time:
                    frame_with_landmarks = apply_landmarks(frame, mp_holistic, target_width, target_height)
                    frame_with_landmarks_list.append(frame_with_landmarks)
                    output_path = os.path.join(output_subfolder, f"frame{frame_count}.jpg")
                    cv2.imwrite(output_path, frame_with_landmarks)

                frame_count += 1

                if current_time >= end_time:
                    break

            cap.release()

            processed_videos.append((video_name, frame_with_landmarks_list))

    mp_holistic.close()

    return processed_videos


def carica_dati_per_rete_bayesiana(file_csv):
    df = pd.read_csv(file_csv)
    emotions = ['Happy', 'Sad', 'Angry', 'Neutral']

    # Creare un dizionario con i nomi dei video come chiavi e le percentuali di emozioni come valori
    dati_video = {}
    for _, row in df.iterrows():
        nome_video, emozione, percentuale = row['Nome Video'], row['Emozione'], float(row['Percentuale'])
        
        # Separare il nome del video e l'emozione
        video_name, emotion_type = nome_video.split('-')
        
        # Aggiungere le percentuali per ogni emozione
        if video_name not in dati_video:
            dati_video[video_name] = {emozione: 0 for emozione in emotions}
        dati_video[video_name][emotion_type] = percentuale

    # Creare le features di input e l'etichetta
    X = []
    y = []
    for video, percentuali in dati_video.items():
        X.append([percentuali[em] for em in emotions])  # Aggiungi percentuali come feature
        label = max(percentuali, key=percentuali.get)  # Etichetta predominante basata sulla percentuale maggiore
        y.append(label)

    # Codifica delle etichette
    le = LabelEncoder()
    y = le.fit_transform(y)

    return X, y, emotions

def carica_dati_per_rete_bayesiana(file_csv):
    df = pd.read_csv(file_csv)
    emotions = ['Happy', 'Sad', 'Angry', 'Neutral']

    # Creare un dizionario con i nomi dei video come chiavi e le percentuali di emozioni come valori
    dati_video = {}
    for _, row in df.iterrows():
        nome_video, emozione = row['Nome Video'], row['Emozione']
        
        # Separare il nome del video e l'emozione
        video_name, emotion_type = nome_video.split('-')
        
        # Aggiungere le percentuali per ogni emozione
        if video_name not in dati_video:
            dati_video[video_name] = {em: 0 for em in emotions}
        dati_video[video_name][emotion_type] = emozione

    # Creare le features di input e l'etichetta
    X = []
    y = []
    for video, percentuali in dati_video.items():
        X.append([percentuali[em] for em in emotions])  # Aggiungi percentuali come feature
        label = max(percentuali, key=percentuali.get)  # Etichetta predominante basata sulla percentuale maggiore
        y.append(label)

    # Codifica delle etichette
    le = LabelEncoder()
    y = le.fit_transform(y)

    return X, y, emotions
def bayesian_model_sklearn(file_csv):
    # Carica i dati
    X, y, emotions = carica_dati_per_rete_bayesiana(file_csv)

    # Split dei dati in train e test
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)
    from imblearn.over_sampling import SMOTE
    # Applicare SMOTE per gestire il bilanciamento delle classi (se necessario)
    smote = SMOTE(random_state=42, k_neighbors=3)
    X_resampled, y_resampled = smote.fit_resample(X_train, y_train)

    # Crea e addestra il modello Naive Bayes
    model = GaussianNB()
    model.fit(X_resampled, y_resampled)
    
    # Previsioni sui dati di test
    y_pred = model.predict(X_test)

    # Converti le etichette in nomi di emozioni
    y_pred_classes = [emotions[label] for label in y_pred]
    y_test_classes = [emotions[label] for label in y_test]

    # Calcola l'accuratezza
    accuracy = accuracy_score(y_test_classes, y_pred_classes)
    print(f"Accuratezza: {accuracy:.2f}")

    # Stampa il classification report
    print("Classification Report:")
    print(classification_report(y_test_classes, y_pred_classes, zero_division=0))

    # Stampa la confusion matrix
    print("Confusion Matrix:")
    print(confusion_matrix(y_test_classes, y_pred_classes))

    # Esegui cross-validation per testare la stabilità del modello
    cv_scores = cross_val_score(model, X_resampled, y_resampled, cv=5)
    print(f"Cross-validation scores: {cv_scores}")
    print(f"Media della cross-validation: {np.mean(cv_scores):.2f}")
    
    # Verifica la presenza di overfitting confrontando accuracy_train vs accuracy_test
    y_train_pred = model.predict(X_resampled)
    y_train_pred_classes = [emotions[label] for label in y_train_pred]
    accuracy_train = accuracy_score(y_resampled, y_train_pred_classes)
    print(f"Accuratezza sui dati di addestramento: {accuracy_train:.2f}")

    # Controllo dell'overfitting: confrontiamo accuracy_train e accuracy_test
    if accuracy_train > accuracy:
        print("Attenzione: Potenziale overfitting rilevato!")
    else:
        print("Il modello sembra non soffrire di overfitting.")


"""
    Funzione **get_folder_names** per ottenere i nomi delle sottocartelle in una cartella.
    Prende in input il percorso della cartella.
    Restituisce una lista di nomi di sottocartelle.
"""


def get_folder_names(output_folder):
    if os.path.isdir(output_folder):
        nomi_cartelle = [nome for nome in os.listdir(output_folder) if os.path.isdir(os.path.join(output_folder, nome))]
        return nomi_cartelle
    else:
        print('Il percorso specificato non corrisponde a una cartella.')
        return []


def main():
    # Crea un parser per gli argomenti da riga di comando
    parser = argparse.ArgumentParser(
        description='Parser per definire gli argomenti da riga di comando (il file .xlsx e la cartella dei video).')
    parser.add_argument('--file', type=str, help='Il percorso del file Excel da processare.')
    parser.add_argument('--video-folder', type=str, help='La cartella contenente i video da elaborare.')

    # Analizza gli argomenti da riga di comando
    args = parser.parse_args()

    # Se l'utente ha fornito un percorso di file, usalo. Altrimenti, usa il percorso di default.
    if args.file:
        file_excel = args.file
    else:
        print(
            "\nStai per utilizzare il percorso di default per il file Excel: 'datasets_xlsx/Responses_train+validation.xlsx'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print(
                "\nPer specificare un percorso di file diverso, esegui lo script con l'opzione \033[91m--file\033[0m, come segue:")
            print(
                "      \033[91mpython\033[0m test_validation_model.py \033[91m--file\033[0m /percorso/del/tuo/file.xlsx\n")
            exit()
        file_excel = '../datasets_xlsx/Responses_train+validation.xlsx'

    risultato = calcola_media_colonne(file_excel)

    # Controllo se è stata fornita una cartella dei video diversa tramite l'argomento --video-folder
    if args.video_folder:
        input_folder = args.video_folder
    elif os.path.exists("./videos"):
        print("\nStai per utilizzare il percorso di default per la cartella dei video: './videos'")
        print("Vuoi continuare? [y/n]")
        response = input()
        if response.lower() != 'y':
            print(
                "\nPer specificare un percorso della cartella dei video diverso, esegui lo script con l'opzione \033[91m--video-folder\033[0m, come segue:")
            print(
                "      \033[91mpython\033[0m test_validation_model.py \033[91m--video-folder\033[0m /percorso/della/tua/cartellaDeiVideo\n")
            exit()
        else:
            input_folder = './videos'
            print("Continuazione dello script in corso...")
    else:
        print(
            "\nÈ necessario specificare un percorso della cartella dei video diverso, esegui lo script con l'opzione \033[91m--video-folder\033[0m, come segue:")
        print(
            "      \033[91mpython\033[0m test_validation_model.py \033[91m--video-folder\033[0m /percorso/della/tua/cartellaDeiVideo\n")
        print("Oppure importa una cartella 'videos' con all'interno i video di tuo interesse nella workspace corrente.")
        exit()
    file_csv = './dativideo.csv'
    with open(file_csv, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['Nome Video', 'Emozione'])

        ultimo_nome_video = ''
        for nome_video, etichetta, percentuale in risultato:
            # Rimuove il prefisso e formatta nome ed etichetta correttamente
            if nome_video.startswith("VID_RGB_000"):
                nome_video = nome_video.replace("VID_RGB_000", "").strip()

            # Scrivi la riga corretta
            writer.writerow([f"{nome_video}{etichetta}", percentuale])
            print()

    print("Dati salvati correttamente nel file CSV:", file_csv)

    output_folder = './outputframe'
    target_width = 180
    target_height = 320

    if os.path.exists("./outputframe") | os.path.exists("./logs") | os.path.exists("emotion_lstm_model.h5"):
        print(
            "\nSiccome esistono rimasugli di vecchie esecuzioni dello script, è necessario cancellarle per assicurarsi che l'esecuzione avvenga senza intoppi.")
        print(
            "Assicurati che questi file sensibili non ti siano utili (\033[91m./outputframe\033[0m - \033[91m./logs\033[0m - \033[91memotion_lstm_model.h5\033[0m)")
        print("Procedo alla cancellazione? [y/n]")
        response = input()
        if response.lower() != 'y':
            print("\nÈ necessario che questi file vengano spostati da questa workspace per il corretto funzionamento.")
            print("Riavvia lo script una volta che questi file siano stati spostati (o cancellati) correttamente.")
            exit()
        else:
            # Cancella la cartella 'outputframe' e tutto il suo contenuto nel caso siano rimasti rimasugli da vecchie esecuzioni
            if os.path.exists("./outputframe"):
                shutil.rmtree('./outputframe')
            if os.path.exists("./logs"):
                shutil.rmtree('./logs')
            if os.path.exists("emotion_lstm_model.h5"):
                os.remove('emotion_lstm_model.h5')
            print("\nCancellazione dei file avvenuta con successo.")
            print("Esecuzione dello script in corso...\n\n")

    nomi_cartelle = get_folder_names(output_folder)
    # Chiamata alla funzione con il file CSV
    bayesian_model_sklearn('./dativideo.csv')


if __name__ == '__main__':
    main()